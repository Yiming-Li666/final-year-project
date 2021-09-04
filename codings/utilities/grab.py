import requests
from bs4 import BeautifulSoup
import json
import openpyxl as xl
import sys
sys.path.append("..") 
from config import Conf
import os

# load the configuration file
conf = Conf("../config/config.json")

cate_path = os.path.join(conf["symptom_dataset_path"], conf["cate_path"])
sub_path = os.path.join(conf["symptom_dataset_path"], conf["sub_path"])
sub2_path = os.path.join(conf["symptom_dataset_path"], conf["sub2_path"])

orig_url = conf["Host_url"]

this = conf["main_category"]
sheet_index = conf["sheet_index"]

this_url = orig_url + "/jibing/"+ this + "/list.htm"
headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "br, gzip, deflate",
    "Accept-Language": "zh-cn",
    "Cache-Control": "no-cache,must-ridate",
    "Cookie": "Hm_lpvt_dfa5478034171cc641b1639b2a5b717d=1598200266; Hm_lvt_dfa5478034171cc641b1639b2a5b717d=1597731311; _ga=GA1.2.1108476298.1597731311; _gid=GA1.2.1182171163.1598199191; CNZZDATA-FE=CNZZDATA-FE; CNZZDATA1256706712=1788341136-1597728253-%7C1598196323; CNZZDATA1914877=cnzz_eid%3D85184446-1597731380-https%253A%252F%252Fwww.haodf.com%252F%26ntime%3D1598199513; UM_distinctid=17400350ad5252-0e6939f84153708-49183500-13c680-17400350ad6a79; g=HDF.164.5f3b71f80ce7f; __jsluid_s=59f72e54e40158955640e5cc1d739ac8",
    "Host": "www.haodf.com",
    "Proxy-Connection": "keep-alive",
    "Referer": "https://www.haodf.com/jibing/xiaoerke/list.htm",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.2 Safari/605.1.15",
    "Content-Type": "text/html; charset=gbk",
    "Transfer-Encoding": "Identity"
}

cate_seg = []
sub_seg = []
sub2_seg = []

def write_to_xsl(sub_list,sub_l):
    print(conf["symptom_dataset_path"])
    wb_path = os.path.join(conf["symptom_dataset_path"], conf["workbook"])
    print(wb_path)
    wb = xl.load_workbook(wb_path)
    wb._active_sheet_index = sheet_index
    ws = wb.active
    for i in range(len(sub_list)):
        ws.cell(row = i+3, column = 2).value = sub_list[i]
    for i in range(len(sub_l)):
        ws.cell(row = i+3, column = 5).value = sub_l[i]
    new_wb_path = os.path.join(conf["symptom_dataset_path"], conf["new_workbook"])
    wb.save(new_wb_path)

# get topli and common disease list
def get_sub(sub_seg,sub2_seg):
    a_count = 0
    b_count = 0
    binary_c = 0
    start_index = sub2_seg.index(sub_seg[0])
    del sub2_seg[:start_index]
    del sub_seg[-1]
    for i in range(len(sub2_seg)):
        if sub_seg[a_count] == " ":
            del sub_seg[a_count]
        # print(a_count)
        # print(len(sub_seg))
        # print(b_count)
        # print(len(sub2_seg))
        if sub_seg[a_count] == sub2_seg[b_count]:
            if binary_c == 0:
                binary_c = 1
                sub2_seg.insert(b_count,'')
                b_count += 1
            if a_count < len(sub_seg) - 1:
                a_count += 1
            b_count += 1
        else:
            if binary_c == 1:
                binary_c = 0
            b_count += 1
    return sub2_seg


# get common disease list
def get_sub_sub(sub_seg,sub2_seg):
    count = 0
    start_index = sub2_seg.index(sub_seg[0])
    del sub2_seg[:start_index]

    for i in range(len(sub2_seg)):
        if len(sub_seg) == 0:
            break
        if sub_seg[0] == " ":
            del sub_seg[0]
        if len(sub_seg) == 0:
            break
        if sub_seg[0] == sub2_seg[count]:
            del sub_seg[0]
            del sub2_seg[count]
            if count == 0:
                sub2_seg.insert(count,'')
                count += 1
            else: 
                if sub2_seg[count-1] != '':
                    sub2_seg.insert(count,'')
                    count += 1
        else:
            count += 1
    return sub2_seg

def main():
    res = requests.get(this_url, headers=headers)
    data = res.text
    data = data.replace("gb2312","utf-8")

    f = open(this+".html", 'w')
    f.write(data)
    f.close()

    soup = BeautifulSoup(data, 'lxml')
    cts = soup.find_all("div", class_="ct")
    for ct in cts:
        m_title_greens = ct.find_all("div", class_="m_title_green")
        for m_title_green in m_title_greens:
            cate_seg.append(m_title_green.span.text)
    m_ctt_greens = soup.find_all("div", class_="m_ctt_green")
    for m_ctt_green in m_ctt_greens:
        toplis = m_ctt_green.find_all("li", class_="topli")
        for topli in toplis:
            sub_seg.append(topli.a.text)
        sub_seg.append(" ")
    print("This page finished!")
    sub = sub_seg
    lis = soup.find_all("li")
    for li in lis:
        a_s = li.find_all("a")
        for a in a_s:
            sub2_seg.append(a.text)
    print("This page finished!")
    
    # # 写入 json 文件中
    # json_str = json.dumps(cate_seg, ensure_ascii=False, indent=4, separators=(',', ':'))
    # json_str2 = json.dumps(sub_seg, ensure_ascii=False, indent=4, separators=(',', ':'))
    # json_str3 = json.dumps(sub2_seg, ensure_ascii=False, indent=4, separators=(',', ':'))
    # f = open(cate_path, 'w', encoding = "utf-8")
    # f.write(json_str)
    # f.close()
    # f = open(sub_path, 'w', encoding = "utf-8")
    # f.write(json_str2)
    # f.close()
    # f = open(sub2_path, 'w', encoding = "utf-8")
    # f.write(json_str3)
    # f.close()
    # print("Json-write finished")

    sub_l = get_sub(sub_seg,sub2_seg)
    #sub_sub = get_sub_sub(sub_seg,sub2_seg)

    write_to_xsl(cate_seg,sub_l)

if __name__ == "__main__":
    main()